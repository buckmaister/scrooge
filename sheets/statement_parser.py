import pandas as pd
from openpyxl import load_workbook
from models.revolut_transaction import TransactionRevolut
from models.unicredit_transaction import TransactionUnicredit
from datetime import datetime


class StatementParser:
    """
    The StatementParser class is designed to read and process bank statements from different formats (CSV, Excel)
    and bank types (Revolut, Unicredit). It transforms raw data into a standardized list of transactions
    with a uniform structure, ready for further processing or storage.

    Attributes:
    - file_path (str): Path to the statement file to be parsed.
    - file_format (str): Indicates the bank type of the statement ('revolut' or 'unicredit'). This is used to
      determine the specific parsing and processing logic to apply.

    Methods:
    - read_data(): Reads the statement file based on its extension (CSV or Excel) and processes the data.
    - _read_excel_to_list(): Helper method to read an Excel file and convert its content into a list of rows.
    - _process_data(data): Processes the raw data rows into a standardized list of transaction objects.
    - _serialize_transaction(transaction): Converts a transaction object into a list of values, serializing
      datetime objects into ISO format strings when necessary.

    Raises:
    - ValueError: If the 'file_format' is not recognized as a supported bank type or if the file is not in an
      expected format (CSV or Excel), or if an unsupported transaction type is encountered during serialization.
    """
    def __init__(self, file_path, file_format):
        self.file_path = file_path
        self.file_format = file_format
        if self.file_format not in ['revolut', 'unicredit']:
            raise ValueError("Unexpected bank type")

    def read_data(self):
        if self.file_path.endswith('.csv'):
            data = pd.read_csv(self.file_path)
        elif self.file_path.endswith('.xlsx'):
            data = self._read_excel_to_list()
        else:
            raise ValueError("The file must be in CSV or Excel format.")
        return self._process_data(data)

    def _read_excel_to_list(self):
        wb = load_workbook(filename=self.file_path)
        sheet = wb.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def _process_data(self, data):
        transactions = []
        for row in data:
            if self.file_format == 'revolut':
                transaction = TransactionRevolut(*row)
            elif self.file_format == 'unicredit':
                transaction = TransactionUnicredit(*row)
            transactions.append(self._serialize_transaction(transaction))
        return transactions

    def _serialize_transaction(self, transaction):
        if isinstance(transaction, TransactionRevolut):
            return [
                transaction.started_date.isoformat() if isinstance(transaction.started_date,
                                                                   datetime) else transaction.started_date,
                transaction.completed_date.isoformat() if isinstance(transaction.completed_date,
                                                                     datetime) else transaction.completed_date,
                transaction.type,
                transaction.product,
                transaction.description,
                transaction.amount,
                transaction.fee,
                transaction.currency,
                transaction.state,
                transaction.balance
            ]
        elif isinstance(transaction, TransactionUnicredit):
            return [
                transaction.data.isoformat() if isinstance(transaction.data, datetime) else transaction.data,
                transaction.description,
                transaction.importo,
                transaction.currency
            ]
        else:
            raise ValueError("Transaction type not supported")
